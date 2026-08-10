from fastapi import FastAPI, Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from copy import deepcopy
import io
import logging
from datetime import datetime

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

app = FastAPI(title="AB Gym Planner API", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "http://localhost:5174",
        "http://127.0.0.1:5174",
        "http://localhost:3000",
        "https://planificacion-rutina.onrender.com",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/health")
async def health_check():
    return {
        "status": "healthy",
        "service": "AB Gym Planner API",
        "timestamp": datetime.now().isoformat(),
        "version": "2.0.0"
    }

@app.get("/")
async def root():
    return {
        "message": "AB Gym Planner API funcionando",
        "endpoints": {
            "health": "/health",
            "generate_xlsx": "/generate-xlsx (POST)"
        }
    }

# --- MODELOS ---
class Ejercicio(BaseModel):
    nombre: str
    series: str
    reps: str
    peso: Optional[str] = ""
    intensidad: str
    pausa: str
    tempo: str
    rir: str

class DiaEntrenamiento(BaseModel):
    nombre_dia: str
    grupo_muscular: str
    ejercicios: List[Ejercicio]

class Planificacion(BaseModel):
    nombre_cliente: str
    objetivo: str
    nivel: str
    frecuencia: str
    dias: List[DiaEntrenamiento]

# --- CONSTANTES DE LA PLANTILLA ---
TEMPLATE_PATH = "PLANTILLA_EXCEL.xlsx"
BLOCK_SIZE = 14          # filas por bloque de día
# Filas de inicio de cada bloque pre-hecho (1-indexed)
BLOCK_STARTS = [2, 16, 30, 44, 58]
MAX_EXERCISES = 11       # filas de ejercicios por bloque

# Columnas (1-indexed)
COL_EJERCICIO = 2    # B
COL_TIPO = 3         # C  (tempo)
COL_INTENSIDAD = 4   # D
COL_PAUSA_GEN = 5    # E  (Tiemp.Desc)
# Columna PESO de cada semana; SERIES=+1, REPS=+2, DESCANSO=+3, RIR=+4
SEMANA_PESO_COLS = [7, 13, 19, 25]  # G, M, S, Y


def copy_block_style(ws, src_start: int, dst_start: int):
    """Copia el formato de un bloque existente a uno nuevo."""
    row_offset = dst_start - src_start

    # Alturas de fila
    for r in range(BLOCK_SIZE):
        src_r = src_start + r
        dst_r = dst_start + r
        if src_r in ws.row_dimensions:
            ws.row_dimensions[dst_r].height = ws.row_dimensions[src_r].height

    # Celdas fusionadas
    merges = [
        (mc.min_row, mc.min_col, mc.max_row, mc.max_col)
        for mc in ws.merged_cells.ranges
        if src_start <= mc.min_row <= src_start + BLOCK_SIZE - 1
    ]
    for min_row, min_col, max_row, max_col in merges:
        ws.merge_cells(
            start_row=min_row + row_offset,
            start_column=min_col,
            end_row=max_row + row_offset,
            end_column=max_col,
        )

    # Estilos y valores de cabecera
    for r in range(BLOCK_SIZE):
        for c in range(1, 30):
            src = ws.cell(row=src_start + r, column=c)
            dst = ws.cell(row=dst_start + r, column=c)
            if isinstance(src, MergedCell) or isinstance(dst, MergedCell):
                continue
            if src.font:      dst.font      = deepcopy(src.font)
            if src.fill:      dst.fill      = deepcopy(src.fill)
            if src.border:    dst.border    = deepcopy(src.border)
            if src.alignment: dst.alignment = deepcopy(src.alignment)
            dst.number_format = src.number_format
            # Copiar texto de las filas de semana y cabecera (offsets 1 y 2)
            if r in (1, 2):
                dst.value = src.value


def fill_day_block(ws, block_start: int, dia: DiaEntrenamiento):
    """Rellena un bloque de día con los datos del plan."""
    title = f"{dia.nombre_dia.upper()} | {dia.grupo_muscular.upper()}"
    ws.cell(row=block_start, column=COL_EJERCICIO).value = title

    for i, ej in enumerate(dia.ejercicios[:MAX_EXERCISES]):
        row = block_start + 3 + i
        ws.cell(row=row, column=COL_EJERCICIO).value   = ej.nombre
        ws.cell(row=row, column=COL_TIPO).value        = ej.tempo or ""
        ws.cell(row=row, column=COL_INTENSIDAD).value  = ej.intensidad or ""
        ws.cell(row=row, column=COL_PAUSA_GEN).value   = ej.pausa or ""

        peso = ej.peso if ej.peso else "-"
        for sc in SEMANA_PESO_COLS:
            ws.cell(row=row, column=sc).value     = peso
            ws.cell(row=row, column=sc + 1).value = ej.series or ""
            ws.cell(row=row, column=sc + 2).value = ej.reps or ""
            ws.cell(row=row, column=sc + 3).value = ej.pausa or ""
            ws.cell(row=row, column=sc + 4).value = ej.rir or ""


def create_xlsx(plan: Planificacion) -> bytes:
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Hoja1"]

    for day_idx, dia in enumerate(plan.dias):
        if day_idx < len(BLOCK_STARTS):
            block_start = BLOCK_STARTS[day_idx]
        else:
            # Crear bloque extra copiando el formato del último bloque de la plantilla
            extra = day_idx - len(BLOCK_STARTS) + 1
            block_start = BLOCK_STARTS[-1] + extra * BLOCK_SIZE
            copy_block_style(ws, BLOCK_STARTS[-1], block_start)

        fill_day_block(ws, block_start, dia)

    # Eliminar los bloques de la plantilla que no fueron utilizados
    n_days = len(plan.dias)
    if n_days < len(BLOCK_STARTS):
        first_unused_row = BLOCK_STARTS[n_days]
        rows_to_delete = (len(BLOCK_STARTS) - n_days) * BLOCK_SIZE
        ws.delete_rows(first_unused_row, rows_to_delete)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@app.post("/generate-xlsx")
async def generate_xlsx_endpoint(plan: Planificacion):
    logger.info(f"📋 Generando Excel para: {plan.nombre_cliente} — {len(plan.dias)} días")
    try:
        xlsx_bytes = create_xlsx(plan)
        safe_name = plan.nombre_cliente.replace(" ", "_")
        logger.info(f"✅ Excel generado ({len(xlsx_bytes)} bytes)")
        return Response(
            content=xlsx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="planificacion_{safe_name}.xlsx"'
            },
        )
    except Exception as e:
        import traceback
        logger.error(f"❌ Error al generar Excel: {e}\n{traceback.format_exc()}")
        return Response(content=f"Error interno: {str(e)}", status_code=500)
